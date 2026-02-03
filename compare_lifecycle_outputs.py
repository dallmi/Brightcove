#!/usr/bin/env python3
"""
Lifecycle Output Comparison Tool

Compares life_cycle_mgmt output from UnifiedPipeline (Excel) against
Harper CSV files to validate consistency between the two systems.

Generates an Excel report with multiple tabs:
- Summary: High-level overview
- File Comparison: Per-account statistics
- Column Analysis: Column differences between systems
- Missing in Unified: Video IDs only in Harper
- Missing in Harper: Video IDs only in Unified
- Value Mismatches: Field-by-field differences for matching videos
- dt_last_viewed Check: Specific validation of this key field

Usage:
    python compare_lifecycle_outputs.py [options]

Examples:
    python compare_lifecycle_outputs.py
    python compare_lifecycle_outputs.py --output report.xlsx
    python compare_lifecycle_outputs.py --harper-dir /path/to/harper --unified-dir /path/to/unified
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# Default paths
DEFAULT_HARPER_DIR = Path("P:/IMPORTANT/Projects/brightcove_ori/Harper/csv")
DEFAULT_UNIFIED_DIR = Path("P:/IMPORTANT/Projects/brightcove_ori/UnifiedPipeline/output/life_cycle_mgmt")

# Account name mapping between Harper and UnifiedPipeline
ACCOUNT_MAPPING = {
    "internet": "Internet",
    "intranet": "Intranet",
    "neo": "neo",
    "research": "research",
    "research_internal": "research_internal",
    "impact": "impact",
    "circleone": "circleone",
    "digital_networks_events": "digital_networks_events",
    "fa_web": "fa_web",
    "sumi_trust": "SuMiTrust",
    "sumitrust": "SuMiTrust",
    "myway": "MyWay",
}

# Columns to ignore in value comparison (too complex or expected to differ)
IGNORE_COLUMNS_IN_VALUE_COMPARE = {
    "images",  # Complex nested JSON
    "cue_points",  # Complex nested JSON
    "text_tracks",  # Complex nested JSON
    "transcripts",  # Complex nested JSON
    "geo",  # Complex nested JSON
    "schedule",  # Complex nested JSON
    "sharing",  # Complex nested JSON
    "link",  # Complex nested JSON
}

# Styles for Excel formatting
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def normalize_account_name(name: str) -> str:
    """Normalize account name for comparison."""
    name = name.lower()
    name = name.replace("_cms_metadata", "").replace("_cms", "")
    name = name.replace(".csv", "").replace(".xlsx", "")
    return name.strip()


def normalize_column_name(col: str) -> str:
    """Normalize column names for comparison."""
    col = col.lower().strip()
    if col.startswith("cf_"):
        col = col[3:]
    return col


def normalize_value(val, ignore_case: bool = True, ignore_whitespace: bool = True) -> str:
    """Normalize a value for comparison."""
    if pd.isna(val) or val is None:
        return ""

    val_str = str(val).strip() if ignore_whitespace else str(val)

    if ignore_case:
        val_str = val_str.lower()

    # Normalize common variations
    if val_str in ("none", "null", "nan", "nat"):
        return ""

    return val_str


def normalize_tags(tags_str: str) -> set:
    """Normalize tags to a set for order-independent comparison."""
    if not tags_str or pd.isna(tags_str):
        return set()
    return set(t.strip().lower() for t in str(tags_str).split(",") if t.strip())


def find_harper_files(harper_dir: Path) -> dict[str, Path]:
    """Find Harper CSV files."""
    files = {}
    if not harper_dir.exists():
        print(f"Warning: Harper directory does not exist: {harper_dir}")
        return files

    for f in harper_dir.glob("*_cms.csv"):
        if "_metadata" in f.name:
            continue
        account = normalize_account_name(f.stem)
        files[account] = f

    if not files:
        for f in harper_dir.glob("*_cms_metadata.csv"):
            account = normalize_account_name(f.stem)
            files[account] = f

    return files


def find_unified_files(unified_dir: Path) -> dict[str, Path]:
    """Find UnifiedPipeline Excel files."""
    files = {}
    if not unified_dir.exists():
        print(f"Warning: UnifiedPipeline directory does not exist: {unified_dir}")
        return files

    month_folders = sorted(
        [d for d in unified_dir.iterdir() if d.is_dir() and "-" in d.name],
        reverse=True
    )

    if not month_folders:
        for f in unified_dir.glob("*_cms.xlsx"):
            account = normalize_account_name(f.stem)
            files[account] = f
        return files

    latest_folder = month_folders[0]
    print(f"Using UnifiedPipeline folder: {latest_folder.name}")

    for f in latest_folder.glob("*_cms.xlsx"):
        account = normalize_account_name(f.stem)
        files[account] = f

    return files


def load_dataframe(file_path: Path) -> pd.DataFrame | None:
    """Load a CSV or Excel file into a DataFrame."""
    try:
        if file_path.suffix == ".csv":
            return pd.read_csv(file_path, low_memory=False, dtype=str)
        elif file_path.suffix == ".xlsx":
            return pd.read_excel(file_path, engine="openpyxl", dtype=str)
        else:
            print(f"Unknown file format: {file_path}")
            return None
    except Exception as e:
        print(f"Error loading {file_path}: {e}")
        return None


def build_column_mapping(harper_cols: list, unified_cols: list) -> dict:
    """Build a mapping between Harper and Unified column names."""
    mapping = {}

    harper_normalized = {normalize_column_name(c): c for c in harper_cols}
    unified_normalized = {normalize_column_name(c): c for c in unified_cols}

    for norm_name, harper_name in harper_normalized.items():
        if norm_name in unified_normalized:
            mapping[harper_name] = unified_normalized[norm_name]

    return mapping


def compare_values(
    harper_val,
    unified_val,
    column_name: str,
    ignore_case: bool = True,
    ignore_whitespace: bool = True,
    ignore_tag_order: bool = True,
) -> tuple[bool, str, str]:
    """
    Compare two values and return (is_match, harper_normalized, unified_normalized).
    """
    # Handle tags specially
    if column_name.lower() == "tags" and ignore_tag_order:
        harper_tags = normalize_tags(harper_val)
        unified_tags = normalize_tags(unified_val)
        return (
            harper_tags == unified_tags,
            ",".join(sorted(harper_tags)),
            ",".join(sorted(unified_tags)),
        )

    harper_norm = normalize_value(harper_val, ignore_case, ignore_whitespace)
    unified_norm = normalize_value(unified_val, ignore_case, ignore_whitespace)

    return harper_norm == unified_norm, harper_norm, unified_norm


def compare_matching_videos(
    df_harper: pd.DataFrame,
    df_unified: pd.DataFrame,
    column_mapping: dict,
    ignore_case: bool = True,
    ignore_whitespace: bool = True,
    ignore_tag_order: bool = True,
    max_mismatches: int = 10000,
) -> list[dict]:
    """Compare values for videos that exist in both datasets."""
    mismatches = []

    # Determine ID columns
    id_col_harper = "id" if "id" in df_harper.columns else "video_id"
    id_col_unified = "id" if "id" in df_unified.columns else "video_id"

    # Create lookup dict for unified
    unified_by_id = df_unified.set_index(df_unified[id_col_unified].astype(str)).to_dict("index")

    # Find matching video IDs
    harper_ids = set(df_harper[id_col_harper].dropna().astype(str))
    unified_ids = set(df_unified[id_col_unified].dropna().astype(str))
    matching_ids = harper_ids & unified_ids

    for video_id in matching_ids:
        if len(mismatches) >= max_mismatches:
            break

        harper_row = df_harper[df_harper[id_col_harper].astype(str) == video_id].iloc[0]
        unified_row = unified_by_id.get(video_id, {})

        if not unified_row:
            continue

        # Compare each mapped column
        for harper_col, unified_col in column_mapping.items():
            # Skip columns we want to ignore
            if normalize_column_name(harper_col) in IGNORE_COLUMNS_IN_VALUE_COMPARE:
                continue

            harper_val = harper_row.get(harper_col, "")
            unified_val = unified_row.get(unified_col, "")

            is_match, harper_norm, unified_norm = compare_values(
                harper_val,
                unified_val,
                harper_col,
                ignore_case,
                ignore_whitespace,
                ignore_tag_order,
            )

            if not is_match:
                mismatches.append({
                    "video_id": video_id,
                    "video_name": harper_row.get("name", ""),
                    "column_harper": harper_col,
                    "column_unified": unified_col,
                    "value_harper": str(harper_val)[:500] if harper_val else "",
                    "value_unified": str(unified_val)[:500] if unified_val else "",
                    "value_harper_normalized": harper_norm[:200],
                    "value_unified_normalized": unified_norm[:200],
                })

    return mismatches


def compare_dt_last_viewed(
    df_harper: pd.DataFrame,
    df_unified: pd.DataFrame,
) -> list[dict]:
    """Specifically compare dt_last_viewed values for all matching videos."""
    results = []

    if "dt_last_viewed" not in df_harper.columns or "dt_last_viewed" not in df_unified.columns:
        return results

    id_col_harper = "id" if "id" in df_harper.columns else "video_id"
    id_col_unified = "id" if "id" in df_unified.columns else "video_id"

    unified_by_id = {
        str(row[id_col_unified]): row.to_dict()
        for _, row in df_unified.iterrows()
        if pd.notna(row[id_col_unified])
    }

    harper_ids = set(df_harper[id_col_harper].dropna().astype(str))
    unified_ids = set(str(k) for k in unified_by_id.keys())
    matching_ids = harper_ids & unified_ids

    for video_id in matching_ids:
        harper_row = df_harper[df_harper[id_col_harper].astype(str) == video_id].iloc[0]
        unified_row = unified_by_id.get(video_id, {})

        harper_lv = harper_row.get("dt_last_viewed", "")
        unified_lv = unified_row.get("dt_last_viewed", "")

        harper_lv_str = str(harper_lv) if pd.notna(harper_lv) else ""
        unified_lv_str = str(unified_lv) if pd.notna(unified_lv) else ""

        # Normalize date formats for comparison
        harper_date = harper_lv_str[:10] if harper_lv_str else ""
        unified_date = unified_lv_str[:10] if unified_lv_str else ""

        status = "MATCH"
        if harper_date != unified_date:
            if not harper_date and unified_date:
                status = "ONLY_IN_UNIFIED"
            elif harper_date and not unified_date:
                status = "ONLY_IN_HARPER"
            else:
                status = "DIFFERENT"

        results.append({
            "video_id": video_id,
            "video_name": harper_row.get("name", ""),
            "harper_dt_last_viewed": harper_lv_str,
            "unified_dt_last_viewed": unified_lv_str,
            "status": status,
        })

    return results


def analyze_account(
    account: str,
    harper_path: Path,
    unified_path: Path,
    ignore_case: bool = True,
    ignore_whitespace: bool = True,
    ignore_tag_order: bool = True,
) -> dict:
    """Perform complete analysis for a single account."""
    result = {
        "account": account,
        "harper_file": str(harper_path) if harper_path else "",
        "unified_file": str(unified_path) if unified_path else "",
        "harper_exists": harper_path.exists() if harper_path else False,
        "unified_exists": unified_path.exists() if unified_path else False,
        "status": "OK",
        "harper_row_count": 0,
        "unified_row_count": 0,
        "harper_column_count": 0,
        "unified_column_count": 0,
        "harper_video_count": 0,
        "unified_video_count": 0,
        "matching_video_count": 0,
        "videos_only_in_harper": [],
        "videos_only_in_unified": [],
        "columns_only_in_harper": [],
        "columns_only_in_unified": [],
        "value_mismatches": [],
        "dt_last_viewed_comparison": [],
        "issues": [],
    }

    if not result["harper_exists"]:
        result["status"] = "HARPER_MISSING"
        result["issues"].append("Harper file does not exist")
        return result

    if not result["unified_exists"]:
        result["status"] = "UNIFIED_MISSING"
        result["issues"].append("UnifiedPipeline file does not exist")
        return result

    # Load dataframes
    df_harper = load_dataframe(harper_path)
    df_unified = load_dataframe(unified_path)

    if df_harper is None:
        result["status"] = "HARPER_LOAD_ERROR"
        result["issues"].append("Failed to load Harper file")
        return result

    if df_unified is None:
        result["status"] = "UNIFIED_LOAD_ERROR"
        result["issues"].append("Failed to load UnifiedPipeline file")
        return result

    # Basic counts
    result["harper_row_count"] = len(df_harper)
    result["unified_row_count"] = len(df_unified)
    result["harper_column_count"] = len(df_harper.columns)
    result["unified_column_count"] = len(df_unified.columns)

    # Column analysis
    harper_cols = set(df_harper.columns)
    unified_cols = set(df_unified.columns)

    harper_cols_normalized = {normalize_column_name(c): c for c in harper_cols}
    unified_cols_normalized = {normalize_column_name(c): c for c in unified_cols}

    harper_only_normalized = set(harper_cols_normalized.keys()) - set(unified_cols_normalized.keys())
    unified_only_normalized = set(unified_cols_normalized.keys()) - set(harper_cols_normalized.keys())

    result["columns_only_in_harper"] = sorted([harper_cols_normalized[c] for c in harper_only_normalized])
    result["columns_only_in_unified"] = sorted([unified_cols_normalized[c] for c in unified_only_normalized])

    # Video ID analysis
    id_col_harper = "id" if "id" in df_harper.columns else "video_id"
    id_col_unified = "id" if "id" in df_unified.columns else "video_id"

    harper_ids = set(df_harper[id_col_harper].dropna().astype(str))
    unified_ids = set(df_unified[id_col_unified].dropna().astype(str))

    result["harper_video_count"] = len(harper_ids)
    result["unified_video_count"] = len(unified_ids)
    result["matching_video_count"] = len(harper_ids & unified_ids)

    ids_only_in_harper = harper_ids - unified_ids
    ids_only_in_unified = unified_ids - harper_ids

    # Get video names for missing IDs
    for vid in ids_only_in_harper:
        row = df_harper[df_harper[id_col_harper].astype(str) == vid]
        name = row.iloc[0].get("name", "") if len(row) > 0 else ""
        result["videos_only_in_harper"].append({"video_id": vid, "name": name})

    for vid in ids_only_in_unified:
        row = df_unified[df_unified[id_col_unified].astype(str) == vid]
        name = row.iloc[0].get("name", "") if len(row) > 0 else ""
        result["videos_only_in_unified"].append({"video_id": vid, "name": name})

    # Build column mapping and compare values
    column_mapping = build_column_mapping(list(df_harper.columns), list(df_unified.columns))

    result["value_mismatches"] = compare_matching_videos(
        df_harper,
        df_unified,
        column_mapping,
        ignore_case,
        ignore_whitespace,
        ignore_tag_order,
    )

    # dt_last_viewed specific comparison
    result["dt_last_viewed_comparison"] = compare_dt_last_viewed(df_harper, df_unified)

    # Determine status and issues
    if result["columns_only_in_harper"]:
        result["issues"].append(f"{len(result['columns_only_in_harper'])} columns only in Harper")
    if result["columns_only_in_unified"]:
        result["issues"].append(f"{len(result['columns_only_in_unified'])} columns only in Unified")
    if ids_only_in_harper:
        result["issues"].append(f"{len(ids_only_in_harper)} videos only in Harper")
    if ids_only_in_unified:
        result["issues"].append(f"{len(ids_only_in_unified)} videos only in Unified")
    if result["value_mismatches"]:
        result["issues"].append(f"{len(result['value_mismatches'])} value mismatches found")

    if result["issues"]:
        result["status"] = "DIFFERENCES_FOUND"

    return result


def create_excel_report(analyses: list[dict], output_path: Path):
    """Generate the Excel report with multiple tabs."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # 1. Summary Tab
    create_summary_tab(wb, analyses)

    # 2. File Comparison Tab
    create_file_comparison_tab(wb, analyses)

    # 3. Column Analysis Tab
    create_column_analysis_tab(wb, analyses)

    # 4. Missing in Unified Tab
    create_missing_in_unified_tab(wb, analyses)

    # 5. Missing in Harper Tab
    create_missing_in_harper_tab(wb, analyses)

    # 6. Value Mismatches Tab
    create_value_mismatches_tab(wb, analyses)

    # 7. dt_last_viewed Check Tab
    create_dt_last_viewed_tab(wb, analyses)

    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")


def style_header_row(ws, num_cols: int):
    """Apply styling to the header row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_adjust_columns(ws):
    """Auto-adjust column widths."""
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


def create_summary_tab(wb: Workbook, analyses: list[dict]):
    """Create the Summary tab with high-level overview."""
    ws = wb.create_sheet("Summary")

    # Title
    ws["A1"] = "Lifecycle Output Comparison Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"] = ""

    # Summary statistics
    total = len(analyses)
    ok_count = sum(1 for a in analyses if a["status"] == "OK")
    diff_count = sum(1 for a in analyses if a["status"] == "DIFFERENCES_FOUND")
    missing_harper = sum(1 for a in analyses if a["status"] == "HARPER_MISSING")
    missing_unified = sum(1 for a in analyses if a["status"] == "UNIFIED_MISSING")

    total_harper_rows = sum(a.get("harper_row_count", 0) for a in analyses)
    total_unified_rows = sum(a.get("unified_row_count", 0) for a in analyses)
    total_harper_videos = sum(a.get("harper_video_count", 0) for a in analyses)
    total_unified_videos = sum(a.get("unified_video_count", 0) for a in analyses)
    total_matching = sum(a.get("matching_video_count", 0) for a in analyses)
    total_mismatches = sum(len(a.get("value_mismatches", [])) for a in analyses)

    summary_data = [
        ["Metric", "Value"],
        ["Total Accounts", total],
        ["Accounts OK (No Differences)", ok_count],
        ["Accounts with Differences", diff_count],
        ["Missing in Harper", missing_harper],
        ["Missing in Unified", missing_unified],
        ["", ""],
        ["Total Rows (Harper)", f"{total_harper_rows:,}"],
        ["Total Rows (Unified)", f"{total_unified_rows:,}"],
        ["Row Difference", f"{total_unified_rows - total_harper_rows:+,}"],
        ["", ""],
        ["Total Videos (Harper)", f"{total_harper_videos:,}"],
        ["Total Videos (Unified)", f"{total_unified_videos:,}"],
        ["Matching Videos", f"{total_matching:,}"],
        ["Videos Only in Harper", f"{total_harper_videos - total_matching:,}"],
        ["Videos Only in Unified", f"{total_unified_videos - total_matching:,}"],
        ["", ""],
        ["Total Value Mismatches", f"{total_mismatches:,}"],
    ]

    for row_idx, row_data in enumerate(summary_data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 4:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            cell.border = THIN_BORDER

    # Status legend
    ws["A25"] = "Status Legend:"
    ws["A25"].font = Font(bold=True)
    ws["A26"] = "OK"
    ws["A26"].fill = OK_FILL
    ws["B26"] = "No differences found"
    ws["A27"] = "DIFFERENCES_FOUND"
    ws["A27"].fill = WARNING_FILL
    ws["B27"] = "Some differences detected"
    ws["A28"] = "MISSING"
    ws["A28"].fill = ERROR_FILL
    ws["B28"] = "File missing in one system"

    auto_adjust_columns(ws)


def create_file_comparison_tab(wb: Workbook, analyses: list[dict]):
    """Create the File Comparison tab with per-account stats."""
    ws = wb.create_sheet("File Comparison")

    headers = [
        "Account", "Status", "Harper Rows", "Unified Rows", "Row Diff",
        "Harper Columns", "Unified Columns", "Harper Videos", "Unified Videos",
        "Matching Videos", "Only in Harper", "Only in Unified", "Value Mismatches"
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))

    for row_idx, analysis in enumerate(sorted(analyses, key=lambda x: x["account"]), start=2):
        data = [
            analysis["account"],
            analysis["status"],
            analysis.get("harper_row_count", 0),
            analysis.get("unified_row_count", 0),
            analysis.get("unified_row_count", 0) - analysis.get("harper_row_count", 0),
            analysis.get("harper_column_count", 0),
            analysis.get("unified_column_count", 0),
            analysis.get("harper_video_count", 0),
            analysis.get("unified_video_count", 0),
            analysis.get("matching_video_count", 0),
            len(analysis.get("videos_only_in_harper", [])),
            len(analysis.get("videos_only_in_unified", [])),
            len(analysis.get("value_mismatches", [])),
        ]

        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER

            # Color status column
            if col == 2:
                if value == "OK":
                    cell.fill = OK_FILL
                elif value == "DIFFERENCES_FOUND":
                    cell.fill = WARNING_FILL
                else:
                    cell.fill = ERROR_FILL

    auto_adjust_columns(ws)


def create_column_analysis_tab(wb: Workbook, analyses: list[dict]):
    """Create the Column Analysis tab showing column differences."""
    ws = wb.create_sheet("Column Analysis")

    headers = ["Account", "Column Name", "Present In", "Notes"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, len(headers))

    row_idx = 2
    for analysis in sorted(analyses, key=lambda x: x["account"]):
        account = analysis["account"]

        for col_name in analysis.get("columns_only_in_harper", []):
            ws.cell(row=row_idx, column=1, value=account)
            ws.cell(row=row_idx, column=2, value=col_name)
            cell = ws.cell(row=row_idx, column=3, value="Harper Only")
            cell.fill = WARNING_FILL
            ws.cell(row=row_idx, column=4, value="Missing in UnifiedPipeline output")
            row_idx += 1

        for col_name in analysis.get("columns_only_in_unified", []):
            ws.cell(row=row_idx, column=1, value=account)
            ws.cell(row=row_idx, column=2, value=col_name)
            cell = ws.cell(row=row_idx, column=3, value="Unified Only")
            cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            ws.cell(row=row_idx, column=4, value="Missing in Harper output")
            row_idx += 1

    if row_idx == 2:
        ws.cell(row=2, column=1, value="No column differences found")
        ws.merge_cells("A2:D2")

    auto_adjust_columns(ws)


def create_missing_in_unified_tab(wb: Workbook, analyses: list[dict]):
    """Create tab listing videos only in Harper (missing from Unified)."""
    ws = wb.create_sheet("Missing in Unified")

    headers = ["Account", "Video ID", "Video Name"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, len(headers))

    row_idx = 2
    for analysis in sorted(analyses, key=lambda x: x["account"]):
        account = analysis["account"]
        for video in analysis.get("videos_only_in_harper", []):
            ws.cell(row=row_idx, column=1, value=account)
            ws.cell(row=row_idx, column=2, value=video["video_id"])
            ws.cell(row=row_idx, column=3, value=video.get("name", ""))
            row_idx += 1

    if row_idx == 2:
        ws.cell(row=2, column=1, value="No videos missing from Unified")
        ws.merge_cells("A2:C2")

    auto_adjust_columns(ws)


def create_missing_in_harper_tab(wb: Workbook, analyses: list[dict]):
    """Create tab listing videos only in Unified (missing from Harper)."""
    ws = wb.create_sheet("Missing in Harper")

    headers = ["Account", "Video ID", "Video Name"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, len(headers))

    row_idx = 2
    for analysis in sorted(analyses, key=lambda x: x["account"]):
        account = analysis["account"]
        for video in analysis.get("videos_only_in_unified", []):
            ws.cell(row=row_idx, column=1, value=account)
            ws.cell(row=row_idx, column=2, value=video["video_id"])
            ws.cell(row=row_idx, column=3, value=video.get("name", ""))
            row_idx += 1

    if row_idx == 2:
        ws.cell(row=2, column=1, value="No videos missing from Harper")
        ws.merge_cells("A2:C2")

    auto_adjust_columns(ws)


def create_value_mismatches_tab(wb: Workbook, analyses: list[dict]):
    """Create tab with detailed value-by-value mismatches."""
    ws = wb.create_sheet("Value Mismatches")

    headers = [
        "Account", "Video ID", "Video Name", "Column (Harper)", "Column (Unified)",
        "Value (Harper)", "Value (Unified)"
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, len(headers))

    row_idx = 2
    for analysis in sorted(analyses, key=lambda x: x["account"]):
        account = analysis["account"]
        for mismatch in analysis.get("value_mismatches", []):
            ws.cell(row=row_idx, column=1, value=account)
            ws.cell(row=row_idx, column=2, value=mismatch["video_id"])
            ws.cell(row=row_idx, column=3, value=mismatch.get("video_name", ""))
            ws.cell(row=row_idx, column=4, value=mismatch["column_harper"])
            ws.cell(row=row_idx, column=5, value=mismatch["column_unified"])
            ws.cell(row=row_idx, column=6, value=mismatch["value_harper"][:500])
            ws.cell(row=row_idx, column=7, value=mismatch["value_unified"][:500])
            row_idx += 1

            if row_idx > 50000:  # Excel row limit safety
                ws.cell(row=row_idx, column=1, value="... truncated (too many mismatches)")
                break

        if row_idx > 50000:
            break

    if row_idx == 2:
        ws.cell(row=2, column=1, value="No value mismatches found")
        ws.merge_cells("A2:G2")

    auto_adjust_columns(ws)


def create_dt_last_viewed_tab(wb: Workbook, analyses: list[dict]):
    """Create tab specifically for dt_last_viewed comparison."""
    ws = wb.create_sheet("dt_last_viewed Check")

    # Summary section
    ws["A1"] = "dt_last_viewed Comparison Summary"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:E1")

    # Calculate summary stats
    total_match = 0
    total_diff = 0
    total_only_harper = 0
    total_only_unified = 0

    for analysis in analyses:
        for comp in analysis.get("dt_last_viewed_comparison", []):
            if comp["status"] == "MATCH":
                total_match += 1
            elif comp["status"] == "DIFFERENT":
                total_diff += 1
            elif comp["status"] == "ONLY_IN_HARPER":
                total_only_harper += 1
            elif comp["status"] == "ONLY_IN_UNIFIED":
                total_only_unified += 1

    summary_data = [
        ["Status", "Count"],
        ["Matching", total_match],
        ["Different Values", total_diff],
        ["Only in Harper", total_only_harper],
        ["Only in Unified", total_only_unified],
    ]

    for row_idx, row_data in enumerate(summary_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            cell.border = THIN_BORDER

    # Detailed section - only show differences
    ws["A10"] = "Detailed Differences (excluding matches)"
    ws["A10"].font = Font(bold=True)

    headers = ["Account", "Video ID", "Video Name", "Harper Value", "Unified Value", "Status"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=11, column=col, value=header)
    style_header_row(ws, len(headers))

    # Adjust header row to row 11
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=11, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row_idx = 12
    for analysis in sorted(analyses, key=lambda x: x["account"]):
        account = analysis["account"]
        for comp in analysis.get("dt_last_viewed_comparison", []):
            if comp["status"] == "MATCH":
                continue  # Skip matches

            ws.cell(row=row_idx, column=1, value=account)
            ws.cell(row=row_idx, column=2, value=comp["video_id"])
            ws.cell(row=row_idx, column=3, value=comp.get("video_name", ""))
            ws.cell(row=row_idx, column=4, value=comp["harper_dt_last_viewed"])
            ws.cell(row=row_idx, column=5, value=comp["unified_dt_last_viewed"])

            status_cell = ws.cell(row=row_idx, column=6, value=comp["status"])
            if comp["status"] == "DIFFERENT":
                status_cell.fill = WARNING_FILL
            elif comp["status"] in ("ONLY_IN_HARPER", "ONLY_IN_UNIFIED"):
                status_cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

            row_idx += 1

            if row_idx > 50000:
                break

        if row_idx > 50000:
            break

    if row_idx == 12:
        ws.cell(row=12, column=1, value="No dt_last_viewed differences found - all values match!")
        ws.merge_cells("A12:F12")
        ws["A12"].fill = OK_FILL

    auto_adjust_columns(ws)


def main():
    parser = argparse.ArgumentParser(
        description="Compare life_cycle_mgmt outputs between Harper and UnifiedPipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python compare_lifecycle_outputs.py
    python compare_lifecycle_outputs.py --output comparison_report.xlsx
    python compare_lifecycle_outputs.py --harper-dir /path/to/harper/csv --unified-dir /path/to/unified
    python compare_lifecycle_outputs.py --case-sensitive --whitespace-sensitive
        """
    )
    parser.add_argument(
        "--harper-dir",
        type=Path,
        default=DEFAULT_HARPER_DIR,
        help=f"Path to Harper CSV directory (default: {DEFAULT_HARPER_DIR})"
    )
    parser.add_argument(
        "--unified-dir",
        type=Path,
        default=DEFAULT_UNIFIED_DIR,
        help=f"Path to UnifiedPipeline life_cycle_mgmt directory (default: {DEFAULT_UNIFIED_DIR})"
    )
    parser.add_argument(
        "--output", "-o",
        type=Path,
        default=Path("lifecycle_comparison_report.xlsx"),
        help="Output Excel file path (default: lifecycle_comparison_report.xlsx)"
    )
    parser.add_argument(
        "--case-sensitive",
        action="store_true",
        help="Make value comparisons case-sensitive"
    )
    parser.add_argument(
        "--whitespace-sensitive",
        action="store_true",
        help="Make value comparisons whitespace-sensitive"
    )
    parser.add_argument(
        "--tag-order-sensitive",
        action="store_true",
        help="Consider tag order when comparing (default: order-independent)"
    )

    args = parser.parse_args()

    print("=" * 60)
    print("Lifecycle Output Comparison Tool")
    print("=" * 60)
    print(f"Harper directory:   {args.harper_dir}")
    print(f"Unified directory:  {args.unified_dir}")
    print(f"Output file:        {args.output}")
    print(f"Case sensitive:     {args.case_sensitive}")
    print(f"Whitespace sensitive: {args.whitespace_sensitive}")
    print(f"Tag order sensitive:  {args.tag_order_sensitive}")
    print("=" * 60)
    print()

    # Find files
    harper_files = find_harper_files(args.harper_dir)
    unified_files = find_unified_files(args.unified_dir)

    print(f"Found {len(harper_files)} Harper files")
    print(f"Found {len(unified_files)} UnifiedPipeline files")
    print()

    if not harper_files and not unified_files:
        print("ERROR: No files found in either directory!")
        print("Please ensure the pipelines have been run and outputs exist.")
        sys.exit(1)

    # Get all unique accounts
    all_accounts = set(harper_files.keys()) | set(unified_files.keys())

    # Also check mapped account names
    for account in list(all_accounts):
        mapped = ACCOUNT_MAPPING.get(account)
        if mapped:
            all_accounts.add(mapped.lower())

    # Analyze each account
    analyses = []
    for account in sorted(all_accounts):
        print(f"Analyzing: {account}...")

        harper_path = harper_files.get(account)

        # Try to find unified file with mapping
        unified_account = ACCOUNT_MAPPING.get(account, account)
        unified_path = (
            unified_files.get(unified_account.lower()) or
            unified_files.get(account) or
            unified_files.get(unified_account)
        )

        if harper_path is None:
            # Try reverse mapping
            for h_key, u_val in ACCOUNT_MAPPING.items():
                if u_val.lower() == account:
                    harper_path = harper_files.get(h_key)
                    break

        # Create placeholder paths for missing files
        if unified_path is None:
            unified_path = args.unified_dir / f"{unified_account}_cms.xlsx"
        if harper_path is None:
            harper_path = args.harper_dir / f"{account}_cms.csv"

        analysis = analyze_account(
            account,
            harper_path,
            unified_path,
            ignore_case=not args.case_sensitive,
            ignore_whitespace=not args.whitespace_sensitive,
            ignore_tag_order=not args.tag_order_sensitive,
        )
        analyses.append(analysis)

    # Generate Excel report
    print()
    print("Generating Excel report...")
    create_excel_report(analyses, args.output)

    # Print summary
    print()
    print("=" * 60)
    print("SUMMARY")
    print("=" * 60)
    ok_count = sum(1 for a in analyses if a["status"] == "OK")
    diff_count = sum(1 for a in analyses if a["status"] == "DIFFERENCES_FOUND")
    print(f"Accounts OK: {ok_count}")
    print(f"Accounts with differences: {diff_count}")
    print(f"Total value mismatches: {sum(len(a.get('value_mismatches', [])) for a in analyses):,}")
    print()
    print(f"Full report saved to: {args.output}")

    # Return exit code
    has_issues = any(a["status"] != "OK" for a in analyses)
    sys.exit(1 if has_issues else 0)


if __name__ == "__main__":
    main()
